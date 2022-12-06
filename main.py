import csv
import math
import re
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import matplotlib.pyplot as plt
import numpy as np
from prettytable import PrettyTable
import doctest

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}


class Vacancy():
    """Класс для представления вакансии

        Atributes:
            name (str): Названиее вакансии
            salary (Salary): Данные о зарплате вакансии
            area_name (str): город места работы
            published_at (str): дата публикации
        """
    def __init__(self, name, salary, area_name, published_at):
        """Выполняет инициализацию объекта Vacancy

        Args:
            name (str): Названиее вакансии
            salary (Salary): Данные о зарплате вакансии
            area_name (str): город места работы
            published_at (str): дата публикации
        """
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at


class Salary():
    """Класс для представления зарплаты

    Atributes:
        salary_from (str or int or float): Нижняя граница оклада
        salary_to (str or int or float): Верхняя граница оклада
        salary_currency (str): Валюта оклада
        salary_gross (str): Брутто оклада
        average_salary (float): Средний оклад
    """
    def __init__(self, salary_from, salary_to, salary_currency, salary_gross=''):
        """Выполняет инициализацию объекта Salary и конвертацию целочисленных полей

        Args:
            salary_from (str or int or float): Нижняя граница оклада
            salary_to (str or int or float): Верхняя граница оклада
            salary_currency (str): Валюта оклада
            salary_gross (str): Брутто оклада
            average_salary (float): Средний оклад
        """
        self.salary_from = float(salary_from) * currency_to_rub[salary_currency]
        self.salary_to = float(salary_to) * currency_to_rub[salary_currency]
        self.salary_gross = salary_gross
        self.salary_currency = salary_currency
        self.average_salary = (self.salary_to + self.salary_from) / 2


class Report():
    """Класс для визуализации статистики

        Atributes:
            profession (str): название профессии для которой сформирована статистика
            all_salary (dict): словарь зависимости годов от средних зарплат
            all_count (dict): словарь зависимости годов от количества вакансий за этот год
            prof_salary (dict): словарь зависимости годов от средних зарплат для выбранной профессии
            prof_count (dict): словарь зависимости годов от количества вакансий за этот год для выбранной профессии
            result_city_salary (dict): словарь зависимости городов от средних зарплат
            result_city_count (dict): словарь зависимости городов от процентной доли зарплат
    """
    def __init__(self, profession, all_salary, all_count, prof_salary, prof_count, result_city_salary,
                 result_city_count):
        """Выполняет инициализацию объекта Report

        Args:
            profession (str): название профессии для которой сформирована статистика
            all_salary (dict): словарь зависимости годов от средних зарплат
            all_count (dict): словарь зависимости годов от количества вакансий за этот год
            prof_salary (dict): словарь зависимости годов от средних зарплат для выбранной профессии                prof_count (dict): словарь зависимости годов от количества вакансий за этот год для выбранной профессии                    result_city_salary (dict): словарь зависимости городов от средних зарплат
            result_city_count (dict): словарь зависимости городов от процентной доли зарплат
        """
        self.profession = profession
        self.all_salary = all_salary
        self.all_count = all_count
        self.prof_salary = prof_salary
        self.prof_count = prof_count
        self.result_city_salary = result_city_salary
        self.result_city_count = result_city_count

    def generate_image(self, name):
        """Формирует изображение с четырьмя графиками по данным статистики

        Args:
            name (str): имя файла, с которым будет сохранено изображение
        """
        fig, axs = plt.subplots(nrows=2, ncols=2)
        font = {'size': 8}

        plt.rc('font', **font)
        years = self.all_salary.keys()
        all_vac = [x[1] for x in self.all_salary.items()]
        prof_vac = [x[1] for x in self.prof_salary.items()]

        x = np.arange(len(years))  # the label locations
        width = 0.35  # the width of the bars
        rects1 = axs[0, 0].bar(x - width / 2, all_vac, width, label='средняя з/п')
        rects2 = axs[0, 0].bar(x + width / 2, prof_vac, width, label=f'з/п {profession}')
        axs[0, 0].set_title('Уровень зарплат по годам')
        axs[0, 0].set_xticks(x, years, rotation='vertical')
        axs[0, 0].yaxis.grid(True)

        years = all_count.keys()
        all_vac = [x[1] for x in self.all_count.items()]
        prof_vac = [x[1] for x in self.prof_count.items()]

        x = np.arange(len(years))  # the label locations
        width = 0.35  # the width of the bars
        rects1 = axs[0, 1].bar(x - width / 2, all_vac, width, label='Количество вакансий')
        rects2 = axs[0, 1].bar(x + width / 2, prof_vac, width, label=f'Количество вакансий {profession}')
        axs[0, 1].set_title('Количесвто вакансий по годам')
        axs[0, 1].set_xticks(x, years, rotation='vertical')
        axs[0, 1].legend(loc='upper left')
        axs[0, 1].yaxis.grid(True)
        cities = [re.sub('-| ', '\n', x) for x in self.result_city_salary.keys()]
        y_pos = np.arange(len(cities))
        salaries = [x[1] for x in self.result_city_salary.items()]
        axs[1, 0].barh(y_pos, salaries, align='center')
        axs[1, 0].set_yticks(y_pos, labels=cities)
        axs[1, 0].invert_yaxis()  # years read top-to-bottom
        axs[1, 0].set_title('Уровень зарплат по городам')
        axs[1, 0].xaxis.grid(True)

        plt.style.use('_mpl-gallery-nogrid')

        # make data
        cities = list(self.result_city_salary.keys())
        cities.append('Другие')
        x = [x[1] for x in self.result_city_count.items()]
        x.append(1 - sum(x))
        colors = plt.get_cmap('tab20')(np.linspace(0.0, 1.0, len(x)))
        axs[1, 1].set_title('Доля вакансий по городам')

        # plot
        axs[1, 1].pie(x, textprops={'fontsize': 6}, colors=colors,
                      labels=cities)

        # all params
        for x in range(2):
            for y in range(2):
                if x == 1 and y == 1:
                    continue
                axs[x, y].tick_params(labelsize=8)
                if x == 0:
                    axs[x, y].legend(loc='upper left')
                if x == 1:
                    axs[1, y].tick_params(labelsize=6)
        fig.set_size_inches(6.40, 4.80, forward=True)
        fig.tight_layout()

        plt.show()
        fig.savefig(name)

    def aligment(self, sheet):
        """Выравнивает ячейки таблицы по содержимому

        Args:
            sheet (): таблица для которой будет произведено выравнивание
        """
        def as_text(val):
            """Преобразует ячейку в строкое представление

            Args:
                val (): ячейка
            Retuns:
                str: ячейка в строковом представлении
            """
            if val is None:
                return ""
            return str(val)

        for column in sheet.columns:
            length = max(len(as_text(cell.value)) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = length + 2

    def stylization(self, sheet, part):
        """Выравнивает ячейки таблицы по содержимому

        Args:
            sheet (): таблица для которой будет произведено стилизация
            part (int): идекс колонки для которой не будет границы сверху и снизу, если не нужно делать отступ, то передать -1
        """
        thin = Side(border_style="thin", color="000000")
        for i in range(1, sheet.max_column + 1):
            sheet.cell(row=1, column=i).font = Font(bold=True)
            for j in range(1, sheet.max_row + 1):
                sheet.cell(row=j, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        if part > 0:
            for j in range(1, sheet.max_row + 1):
                sheet.cell(row=j, column=part).border = Border(left=thin, right=thin)

    def generate_excel(self, name):
        """Формирует exel таблицу по данным статистики

        Args:
            name (str): имя файла, с которым будет сохранена таблица
        """
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Статистика по годам'
        header = ['Год', 'Средняя зарплата', f'Средняя зарплата - {profession}', 'Количество вакансий',
                  f'Количество вакансий - {profession}']
        sheet.append(header)
        years = list(all_salary.keys())
        for i in range(len(years)):
            row = []
            row.append(years[i])
            row.append(list(map(list, all_salary.items()))[i][1])
            row.append(list(map(list, prof_salary.items()))[i][1])
            row.append(list(map(list, all_count.items()))[i][1])
            row.append(list(map(list, all_salary.items()))[i][1])
            sheet.append(row)
        self.aligment(sheet)
        self.stylization(sheet, -1)

        wb.create_sheet('Статистика по городам', 1)
        wb.active = 1
        sheet = wb.active
        header = ['Город', 'Уровень зарплат', '	', 'Город', 'Доля вакансий']
        sheet.append(header)
        cities = list(result_city_salary.keys())
        for i in range(len(cities)):
            row = []
            row.append(cities[i])
            row.append(list(map(list, result_city_salary.items()))[i][1])
            row.append('')
            row.append(cities[i])
            row.append(list(map(list, result_city_count.items()))[i][1])
            sheet.append(row)
        for i in range(1, sheet.max_row + 1):
            sheet.cell(row=i, column=5).number_format = '0.00%'
        self.aligment(sheet)
        self.stylization(sheet, 3)
        wb.save(name)


def clear_string(text):
    """Очищает строку от html тегов и заменяет перенос строки на запятую

    Args:
        text (str): изменяемая строка
    Returns:
        str: очищенная строка
    """
    output = re.sub(r"<.*?>", "", text)
    output = output.strip()
    if output.find("\n") != -1:
        output = ', '.join(output.split('\n'))
    return re.sub("\s+", " ", output)


def csv_reader(file_name):
    """Считывает csv файл

    Args:
        file_name (str): путь до файла
    Returns:
        str[]: данные таблицы
        str[]: заголовки таблицы
    """
    with open(file_name, encoding="utf-8-sig") as file:
        text = csv.reader(file)
        lines = []
        head_line = []
        for line in text:
            lines.append(line)
        if len(lines) > 0:
            head_line = lines.pop(0)
    file.close()
    return lines, head_line


def csv_filer(reader, list_naming):
    """Преобразует данные в коллекцию словарей с данными о вакансииях

    Args:
        reader (str[]): данные таблицы
        list_naming (str[]): заголовки таблицы
    Returns:
        dict[]: коллекция словарей с данными о вакансиях
    """
    vacancies = []
    lines = []
    for line in reader:
        line_without_empty = [x for x in line if x != ""]
        if len(line_without_empty) == len(list_naming):
            lines.append(line_without_empty)
    for line in lines:
        vacancy = {}
        for i in range(len(list_naming)):
            vacancy[list_naming[i]] = clear_string(line[i])
        vacancies.append(vacancy)
    return vacancies


def up_count(key, dict):
    """Выполняет инкремент по заданному ключу

    Args:
        key (str): ключ словаря
        dict (dict): изменяемый словарь
    Returns:
        dict: измененный словарь
    """
    try:
        key = int(key)
    except:
        key = key
    if dict.__contains__(key):
        dict[key] += 1
    else:
        dict[key] = 1
    return dict


def up_salary(key, dict, salary):
    """Выполняет инкремент элемента по заданному ключу и увеличивает зарплату на заданное число

    Args:
        key (str): ключ словаря
        dict (dict): изменяемый словарь
        salary (int): прибавляемая зарплата
    Returns:
        dict: измененный словарь
    """
    try:
        key = int(key)
    except:
        key = key
    if dict.__contains__(key):
        dict[key] = (dict[key][0] + 1, dict[key][1] + salary)
    else:
        dict[key] = (1, salary)
    return dict


def get_average_salary_by_year(dict):
    """Возращает словарь годов по средним зарплатам

    Args:
        dict (dict): словарь годов по общей зарпалате и их количеству
    Returns:
        dict: словарь годов по средним зарплатам
    """
    res = {}
    for key in dict.keys():
        try:
            res[key] = int(math.floor(dict[key][1] / dict[key][0]))
        except:
            res[key] = 0
    return res


def dict_init_salary(key, dict):
    """Инициализирует словарь зарплат по заданному ключу

    Args:
        key (str): ключ словаря
        dict (dict): инициализируемый словарь
    Returns:
        dict: инициализированный словарь
    """
    try:
        key = int(key)
    except:
        key = key
    if not dict.__contains__(key):
        dict[key] = (0, 0)
    return dict


def dict_init_count(key, dict):
    """Инициализирует словарь количеств по заданному ключу

    Args:
        key (str): ключ словаря
        dict (dict): инициализируемый словарь
    Returns:
        dict: инициализированный словарь
    """
    try:
        key = int(key)
    except:
        key = key
    if not dict.__contains__(key):
        dict[key] = 0
    return dict


def get_statistics(list_vacancies, prof, actual_city):
    """Формирует словари со статистикой

    Args:
        list_vacancies (dict[]): данные о вакансиях
        prof (str): профессия по которой будет сформирована статистика
        actual_city (): топ 10 городов по количеству вакансий
    Returns:
        dict: словарь годов по количеству
        dict: словарь годов по количеству заданной профессии
        dict: словарь годов по средней зарплате
        dict: словарь годов по средней зарплате заданной профессии
        dict: словарь городов по средней зарплате
    """
    all_count = {}
    all_salary = {}
    prof_count = {}
    prof_salary = {}
    city_salary = {}

    for vacancy in list_vacancies:
        is_right_prof = prof in vacancy.name
        date = vacancy.published_at[:4]
        prof_count = dict_init_count(date, prof_count)
        prof_salary = dict_init_salary(date, prof_salary)
        all_count = up_count(date, all_count)
        all_salary = up_salary(date, all_salary, vacancy.salary.average_salary)
        if vacancy.area_name in actual_city:
            city_salary = up_salary(vacancy.area_name, city_salary, vacancy.salary.average_salary)
        if is_right_prof:
            prof_count = up_count(date, prof_count)
            prof_salary = up_salary(date, prof_salary, vacancy.salary.average_salary)

    return all_count, prof_count, get_average_salary_by_year(all_salary), get_average_salary_by_year(
        prof_salary), get_average_salary_by_year(city_salary)


def fill_table(vac_data, table):
    """Заполняет таблицу вакансиями

    Args:
        vac_data (dict[]): данные о вакансиях
        table (PrettyTable): заполняемая таблица
    Returns:
        PrettyTable: заполненная таблица
    """
    table.align = "l"
    table.field_names = ['№'] + list(vacancies[0].keys())
    table.max_width = 20
    table.hrules = True
    row = []
    count = 0
    for vacancy in vac_data:
        count += 1
        row.append(str(count))
        for key in vacancy.keys():
            row.append(vacancy[key])
        table.add_row(row)
        row = []
    return table


file_name = 'vacancies.csv'  # input('Введите название файла: ')
profession = 'аналитик'  # input('Введите название профессии: ')
command = input('таблица или ексель или график')
lines, head = csv_reader(file_name)
vacancies = csv_filer(lines, head)

table = PrettyTable()

list_vacancies = []
city_count = {}
for vacancy in vacancies:
    salary = Salary(vacancy["salary_from"], vacancy["salary_to"], vacancy["salary_currency"])
    vacancy_obj = Vacancy(vacancy['name'], salary, vacancy['area_name'], vacancy['published_at'])
    city_count = up_count(vacancy_obj.area_name, city_count)
    list_vacancies.append(vacancy_obj)
table = fill_table(vacancies, table)
actual_city = []
for city in city_count.keys():
    if city_count[city] / len(list_vacancies) >= 0.01:
        actual_city.append(city)

all_count, prof_count, all_salary, prof_salary, city_salary = get_statistics(list_vacancies, profession, actual_city)

sorted_city_count = sorted(city_count, key=city_count.get, reverse=True)
result_city_count = {}

sorted_city_salary = sorted(city_salary, key=city_salary.get, reverse=True)
result_city_salary = {}

for city in sorted_city_salary[:10]:
    result_city_salary[city] = city_salary[city]

for city in sorted_city_count[:10]:
    if city in actual_city:
        share = round(city_count[city] / len(list_vacancies), 4)
        result_city_count[city] = share
rep = Report(profession, all_salary, all_count, prof_salary, prof_count, result_city_salary, result_city_count)
if len(head) == 0:
    print('Пустой файл')
elif command == 'ексель':
    rep.generate_excel('develop.xlsx')
elif command == 'ексель':
    rep.generate_image('graph.png')
elif command == 'таблица':
    print(table)
else:
    print(f'Динамика уровня зарплат по годам: {all_salary}\n'
          f'Динамика количества вакансий по годам: {all_count}\n'
          f'Динамика уровня зарплат по годам для выбранной профессии: {prof_salary}\n'
          f'Динамика количества вакансий по годам для выбранной профессии: {prof_count}\n'
          f'Уровень зарплат по городам (в порядке убывания): {result_city_salary}\n'
          f'Доля вакансий по городам (в порядке убывания): {result_city_count}')
